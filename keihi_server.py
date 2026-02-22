# -*- coding: utf-8 -*-
"""
🧾 KeihiAI - 経費精算自動化サーバー
======================================
起動方法:
1. pip install flask flask-cors pillow openpyxl anthropic
2. ANTHROPIC_API_KEY環境変数を設定（任意・なしでもサンプル動作）
3. python keihi_server.py
4. ブラウザで http://localhost:5001 を開く
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import threading, os, base64, json, re
from pathlib import Path
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder=".")
CORS(app)

status = {"step": 0, "done": False, "error": None, "count": 0, "total": 0, "categories": 0}
UPLOAD_DIR = Path("uploads_temp")
UPLOAD_DIR.mkdir(exist_ok=True)

# カテゴリ自動判定
CATEGORY_RULES = {
    "駐車場": ["パーキング", "駐車", "parking", "コインパーク"],
    "交通費": ["電車", "バス", "タクシー", "新幹線", "乗車", "IC"],
    "飲食費": ["レストラン", "カフェ", "食堂", "居酒屋", "ランチ", "コーヒー", "食事"],
    "宿泊費": ["ホテル", "旅館", "宿", "inn", "hotel"],
    "消耗品": ["コンビニ", "ドラッグ", "文具", "ローソン", "セブン", "ファミマ"],
    "通信費": ["ドコモ", "au", "ソフトバンク", "通信", "インターネット"],
}

def guess_category(text):
    for cat, keywords in CATEGORY_RULES.items():
        for kw in keywords:
            if kw in text:
                return cat
    return "その他"

def extract_amount(text):
    patterns = [
        r'領収額[^\d]*(\d[\d,]+)円',
        r'現金[^\d]*(\d[\d,]+)円',
        r'合計[^\d]*(\d[\d,]+)円',
        r'(\d[\d,]+)円',
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            return int(m.group(1).replace(',', ''))
    return 0

def image_to_base64(path):
    with open(path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode()

def read_receipt_with_claude(image_path):
    """Claude APIで領収書を読み取る（APIキーがある場合）"""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return None

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        img = Image.open(image_path)
        img.save(str(image_path) + "_resized.jpg", "JPEG", quality=85)
        b64 = image_to_base64(str(image_path) + "_resized.jpg")

        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=500,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
                    {"type": "text", "text": "この領収書から以下をJSON形式で抽出してください。{\"店名\": \"\", \"日付\": \"\", \"金額\": 0, \"カテゴリ\": \"\", \"支払方法\": \"\", \"備考\": \"\"}"}
                ]
            }]
        )
        text = response.content[0].text
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            return json.loads(match.group())
    except Exception as e:
        print(f"Claude API error: {e}")
    return None

def fallback_read(image_path, filename):
    """APIなしのフォールバック（ファイル名とサンプルデータで処理）"""
    # 実際のアプリではpytesseract等でOCRする
    return {
        "店名": filename.replace(".jpg", "").replace(".png", "").replace(".pdf", ""),
        "日付": "2025/10/01",
        "金額": 1000,
        "カテゴリ": "その他",
        "支払方法": "現金",
        "備考": "手動確認推奨"
    }



# freee 勘定科目マッピング
FREEE_ACCOUNT_MAP = {
    "駐車場": "旅費交通費",
    "交通費": "旅費交通費",
    "飲食費": "交際費",
    "宿泊費": "旅費交通費",
    "消耗品": "消耗品費",
    "通信費": "通信費",
    "その他": "雑費",
}

def make_freee_csv(receipts, month, applicant):
    import csv
    total = 0
    cats = set()
    with open("expense_report_freee.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        # freee正式ヘッダー
        writer.writerow([
            "発生日","借方勘定科目","借方補助科目","借方税区分","借方金額",
            "貸方勘定科目","貸方補助科目","貸方税区分","貸方金額",
            "摘要","タグ","メモ","決済期日","口座"
        ])
        for r in receipts:
            cat = r.get("カテゴリ","その他")
            account = FREEE_ACCOUNT_MAP.get(cat, "雑費")
            amt = r.get("金額", 0)
            date = r.get("日付","").replace("-","/")
            memo = r.get("店名","") + ("（" + applicant + "）" if applicant else "")
            writer.writerow([
                date, account, "", "課税仕入10%", str(amt),
                "現金", "", "", str(amt),
                memo, "", "", "", ""
            ])
            total += amt
            cats.add(cat)
    return total, len(cats)

def make_csv(receipts, month, applicant):
    import csv
    total = 0
    cats = set()
    with open("expense_report.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["No.", "店名", "日付", "カテゴリ", "金額（円）", "支払方法", "備考"])
        for i, r in enumerate(receipts, 1):
            writer.writerow([i, r.get("店名",""), r.get("日付",""), r.get("カテゴリ",""),
                             r.get("金額",0), r.get("支払方法","現金"), r.get("備考","")])
            total += r.get("金額", 0)
            cats.add(r.get("カテゴリ","その他"))
        writer.writerow(["合計", "", "", "", total, "", ""])
    return total, len(cats)

def make_pdf(receipts, month, applicant):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))

    doc = SimpleDocTemplate("expense_report.pdf", pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title_style = styles["Title"]
    title_style.fontName = "HeiseiKakuGo-W5"
    story.append(Paragraph(f"経費精算書　{month}　{applicant or ''}", title_style))
    story.append(Spacer(1, 16))

    data = [["No.", "店名", "日付", "カテゴリ", "金額（円）", "支払方法"]]
    total = 0
    cats = set()
    for i, r in enumerate(receipts, 1):
        data.append([str(i), r.get("店名",""), r.get("日付",""), r.get("カテゴリ",""),
                     f"{r.get('金額',0):,}", r.get("支払方法","現金")])
        total += r.get("金額", 0)
        cats.add(r.get("カテゴリ","その他"))
    data.append(["合計", "", "", "", f"{total:,}", ""])

    table = Table(data, colWidths=[30, 100, 65, 65, 65, 65])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2D6A4F")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,-1), "HeiseiKakuGo-W5"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#E8F5EE")),
        ("FONTNAME", (0,-1), (-1,-1), "HeiseiKakuGo-W5"),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.white, colors.HexColor("#F4FAF6")]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#DDDDDD")),
    ]))
    story.append(table)
    doc.build(story)
    return total, len(cats)

def make_excel(receipts, month, applicant):
    wb = Workbook()
    ws = wb.active
    ws.title = "経費一覧"

    thin = Side(style="thin", color="DDDDDD")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    # タイトル
    ws.merge_cells("A1:G1")
    ws["A1"].value = f"経費精算書　{month}　申請者：{applicant or '未記入'}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="2D6A4F")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill("solid", start_color="E8F5EE")
    ws.row_dimensions[1].height = 32

    # ヘッダー
    headers = ["No.", "店名", "日付", "カテゴリ", "金額（円）", "支払方法", "備考"]
    widths =  [5,     24,    14,    14,          13,          12,          28]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=i)
        c.value = h
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color="2D6A4F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 22

    # データ
    total = 0
    for i, r in enumerate(receipts, 1):
        row = i + 2
        bg = "FFFFFF" if i % 2 == 1 else "F4FAF6"
        vals = [i, r.get("店名",""), r.get("日付",""), r.get("カテゴリ",""), r.get("金額",0), r.get("支払方法","現金"), r.get("備考","")]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col)
            c.value = val
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
            c.border = bdr
            if col == 5: c.number_format = '#,##0'
        total += r.get("金額", 0)
        ws.row_dimensions[row].height = 20

    # 合計行
    tr = len(receipts) + 3
    ws.merge_cells(f"A{tr}:D{tr}")
    c = ws.cell(row=tr, column=1)
    c.value = "合計"
    c.font = Font(name="Arial", bold=True, size=11)
    c.fill = PatternFill("solid", start_color="E8F5EE")
    c.alignment = Alignment(horizontal="center")
    c.border = bdr

    c = ws.cell(row=tr, column=5)
    c.value = f"=SUM(E3:E{tr-1})"
    c.font = Font(name="Arial", bold=True, size=12, color="2D6A4F")
    c.fill = PatternFill("solid", start_color="E8F5EE")
    c.alignment = Alignment(horizontal="center")
    c.number_format = '#,##0'
    c.border = bdr

    for col in [6,7]:
        ws.cell(row=tr, column=col).fill = PatternFill("solid", start_color="E8F5EE")
        ws.cell(row=tr, column=col).border = bdr
    ws.row_dimensions[tr].height = 26

    # カテゴリ別集計シート
    ws2 = wb.create_sheet("カテゴリ別集計")
    cat_totals = {}
    for r in receipts:
        cat = r.get("カテゴリ","その他")
        cat_totals[cat] = cat_totals.get(cat, 0) + r.get("金額", 0)

    ws2.append(["カテゴリ", "金額（円）"])
    for cat, amt in cat_totals.items():
        ws2.append([cat, amt])

    # 棒グラフ
    chart = BarChart()
    chart.type = "col"
    chart.title = "カテゴリ別経費"
    chart.y_axis.title = "金額（円）"
    chart.style = 10
    chart.width = 16
    chart.height = 10
    data = Reference(ws2, min_col=2, max_col=2, min_row=1, max_row=len(cat_totals)+1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=len(cat_totals)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "52B788"
    ws2.add_chart(chart, "D2")

    output = "expense_report.xlsx"
    wb.save(output)
    return total, len(cat_totals)

def process_files(files_data, month, applicant):
    global status
    try:
        receipts = []

        # Step1: ファイル読み込み
        status["step"] = 1
        saved_paths = []
        for name, data in files_data:
            path = UPLOAD_DIR / name
            path.write_bytes(data)
            saved_paths.append((name, path))

        # Step2: AI文字認識
        status["step"] = 2
        for name, path in saved_paths:
            result = read_receipt_with_claude(path)
            if not result:
                # フォールバック
                result = fallback_read(path, name)
            receipts.append(result)

        # Step3: カテゴリ分類
        status["step"] = 3
        for r in receipts:
            if not r.get("カテゴリ") or r["カテゴリ"] == "その他":
                text = r.get("店名","") + r.get("備考","")
                r["カテゴリ"] = guess_category(text)

        # Step4: 整形
        status["step"] = 4

        # Step5: 出力形式に応じて生成
        status["step"] = 5
        fmt = status.get("format", "excel")
        if fmt == "csv":
            total, cats = make_csv(receipts, month, applicant)
        elif fmt == "freee":
            total, cats = make_freee_csv(receipts, month, applicant)
        elif fmt == "pdf":
            total, cats = make_pdf(receipts, month, applicant)
        else:
            total, cats = make_excel(receipts, month, applicant)

        status.update({"done": True, "count": len(receipts), "total": total, "categories": cats})

    except Exception as e:
        status["error"] = str(e)

@app.route("/")
def index():
    return send_file("keihi_app.html")

@app.route("/analyze", methods=["POST"])
def analyze():
    global status
    status = {"step": 0, "done": False, "error": None, "count": 0, "total": 0, "categories": 0}
    files_data = [(f.filename, f.read()) for f in request.files.getlist("files")]
    month = request.form.get("month", "")
    applicant = request.form.get("applicant", "")
    status["format"] = request.form.get("format", "excel")
    threading.Thread(target=process_files, args=(files_data, month, applicant)).start()
    return jsonify({"ok": True})

@app.route("/status_expense")
def get_status():
    return jsonify(status)

@app.route("/download_expense")
def download():
    fmt = status.get("format", "excel")
    files_map = {
        "excel": ("expense_report.xlsx", "expense_report.xlsx"),
        "csv":   ("expense_report.csv",  "expense_report.csv"),
        "freee": ("expense_report_freee.csv", "freee_import.csv"),
        "pdf":   ("expense_report.pdf",  "expense_report.pdf"),
    }
    filename, dl_name = files_map.get(fmt, files_map["excel"])
    path = Path(filename)
    if path.exists():
        return send_file(str(path), as_attachment=True, download_name=dl_name)
    return jsonify({"error": "ファイルが見つかりません"}), 404

if __name__ == "__main__":
    print("🧾 KeihiAI 起動中...")
    print("👉 ブラウザで http://localhost:5001 を開いてください")
    print("💡 Claude APIキーを設定するとAI読み取りが有効になります")
    print("   例: set ANTHROPIC_API_KEY=sk-ant-...")
    app.run(host="0.0.0.0", debug=False, port=10000)
